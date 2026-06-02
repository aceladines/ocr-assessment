import tempfile
from pathlib import Path
import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.core.config import settings
from app.services.ocr_service import OCRService
from app.services.excel_service import ExcelService

client = TestClient(app)

def test_root_health_check():
    """Verify that the welcome health-check endpoint works and returns correct status metadata."""
    response = client.get("/")
    assert response.status_code == 200
    data = response.json()
    assert data["status"] == "healthy"
    assert data["mock_ocr_enabled"] is True

def test_ocr_response_parsing():
    """Verify that OCRService correctly parses and sanitizes raw OCR response fields."""
    service = OCRService(mock=False)
    
    # Test valid text data
    raw_payload = {
        "page": "1",
        "receipt_no": "OR-12345",
        "doctor_name": "Dr. House",
        "prc_license": "123456",
        "hospital": "Princeton Plainsboro",
        "date": "2026-05-15",
        "patient_name": "John Doe",
        "total_amount": "₱12,345.67 PHP"
    }
    
    sanitized = service._sanitize_parsed_dict(raw_payload)
    assert sanitized["page"] == 1
    assert sanitized["receipt_no"] == "OR-12345"
    assert sanitized["doctor_name"] == "Dr. House"
    assert sanitized["total_amount"] == 12345.67

    # Test error fallback handling for broken float amounts
    raw_payload_bad_amount = {
        "total_amount": "invalid_amount_string"
    }
    sanitized_bad = service._sanitize_parsed_dict(raw_payload_bad_amount)
    assert sanitized_bad["total_amount"] == 0.0

def test_ocr_response_parsing_nested_live():
    """Verify that OCRService correctly parses the live Nanonets layout nested inside result.json.content."""
    service = OCRService(mock=False)
    
    nested_payload = {
        "success": True,
        "message": "Extraction completed successfully",
        "result": {
            "json": {
                "content": {
                    "page": 1,
                    "receipt_no": "OR-70875",
                    "doctor_name": "Dr. Emilio Aguinaldo",
                    "prc_license": "00005",
                    "hospital": "CARDINAL SANTOS MEDICAL CENTER",
                    "date": "2024-04-30",
                    "patient_name": "Bernardo Q. Tolentino",
                    "total_amount": 2050
                }
            }
        }
    }
    
    parsed = service._parse_ocr_response(nested_payload)
    assert parsed["receipt_no"] == "OR-70875"
    assert parsed["doctor_name"] == "Dr. Emilio Aguinaldo"
    assert parsed["total_amount"] == 2050.0

@pytest.mark.asyncio
async def test_excel_sheet_generation():
    """Verify that the ExcelService creates and styles spreadsheet files correctly."""
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = Path(tmpdir)
        excel_service = ExcelService(output_dir=tmp_path)
        
        sample_file = Path("test_invoice.pdf")
        sample_record = {
            "page": 1,
            "receipt_no": "OR-9999",
            "doctor_name": "Dr. Strange",
            "prc_license": "987654",
            "hospital": "Kamar-Taj Clinic",
            "date": "2026-05-28",
            "patient_name": "Stephen Strange",
            "total_amount": 1500.50
        }
        
        # Write the batch (the live code path) which creates and styles the sheet
        await excel_service.write_batch_records(sample_file, [sample_record])

        expected_excel = tmp_path / "test_invoice.xlsx"
        assert expected_excel.exists() is True
        
        # Check workbook structural integrity
        from openpyxl import load_workbook
        wb = load_workbook(expected_excel)
        assert "Extracted Invoices" in wb.sheetnames
        ws = wb["Extracted Invoices"]
        
        # Check headers match
        headers_read = [cell.value for cell in ws[1]]
        assert headers_read[0] == "Page"
        assert headers_read[-1] == "Total Amount (PHP)"
        
        # Check that gridlines are enabled
        assert ws.views.sheetView[0].showGridLines is True
        
        # Check second row data
        row2_read = [cell.value for cell in ws[2]]
        assert row2_read[0] == 1
        assert row2_read[1] == "OR-9999"
        assert row2_read[2] == "Dr. Strange"
        assert row2_read[7] == 1500.50

def test_api_upload_endpoint():
    """Verify that direct uploading of files works and returns valid Excel links."""
    # Create temporary PDF files to upload
    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_1 = Path(tmpdir) / "invoice_test_1.pdf"
        pdf_2 = Path(tmpdir) / "invoice_test_2.pdf"
        
        pdf_1.write_bytes(b"%PDF-1.4 mock content 1")
        pdf_2.write_bytes(b"%PDF-1.4 mock content 2")
        
        with open(pdf_1, "rb") as f1, open(pdf_2, "rb") as f2:
            files_payload = [
                ("files", ("invoice_test_1.pdf", f1, "application/pdf")),
                ("files", ("invoice_test_2.pdf", f2, "application/pdf"))
            ]
            
            response = client.post("/api/v1/extract/upload", files=files_payload)
            
        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["processed_count"] == 2
        assert data["successful_count"] == 2
        assert "excel_download_url" in data
        assert data["excel_download_url"] is not None

def test_api_folder_scan_endpoint(monkeypatch):
    """Verify that scanning a directory path triggers OCR on files and saves spreadsheet."""
    with tempfile.TemporaryDirectory() as tmpdir:
        # Create 3 dummy PDF files in a subdirectory
        pdf_dir = Path(tmpdir) / "medical_invoices"
        pdf_dir.mkdir()

        # Folder scans are confined to ALLOWED_SCAN_DIR; point it at the tmp root.
        monkeypatch.setattr(settings, "ALLOWED_SCAN_DIR", tmpdir)

        (pdf_dir / "invoice_a.pdf").write_bytes(b"%PDF-1.4 mock a")
        (pdf_dir / "invoice_b.pdf").write_bytes(b"%PDF-1.4 mock b")
        (pdf_dir / "ignore_me.txt").write_bytes(b"not a pdf")

        # Post folder scan request
        payload = {"folder_path": str(pdf_dir)}
        response = client.post("/api/v1/extract/folder", json=payload)

        assert response.status_code == 200
        data = response.json()
        assert data["success"] is True
        assert data["processed_count"] == 2
        assert data["successful_count"] == 2
        assert data["failed_count"] == 0
        assert len(data["results"]) == 2


def test_folder_scan_outside_allowlist_rejected(monkeypatch):
    """Folder scans outside ALLOWED_SCAN_DIR must be rejected (no arbitrary FS read)."""
    with tempfile.TemporaryDirectory() as allowed, tempfile.TemporaryDirectory() as outside:
        monkeypatch.setattr(settings, "ALLOWED_SCAN_DIR", allowed)
        (Path(outside) / "secret.pdf").write_bytes(b"%PDF-1.4 secret")

        response = client.post(
            "/api/v1/extract/folder", json={"folder_path": outside}
        )
        assert response.status_code == 400
        assert "permitted scan directory" in response.json()["message"]


def test_download_rejects_path_traversal():
    """The download endpoint must never serve files outside the output dir."""
    # Encoded traversal attempt must not return file contents.
    response = client.get(
        "/api/v1/extract/download/..%2f..%2f..%2f..%2fetc%2fpasswd"
    )
    assert response.status_code in (400, 404)
    assert b"root:" not in response.content

    # A bare, non-existent name resolves cleanly to a 400 (not a 500).
    response = client.get("/api/v1/extract/download/does_not_exist.xlsx")
    assert response.status_code == 400


def test_auth_enforced_when_key_configured(monkeypatch):
    """When API_AUTH_KEY is set, extraction endpoints require a matching header."""
    monkeypatch.setattr(settings, "API_AUTH_KEY", "s3cret")

    # Missing key -> 401
    resp = client.post("/api/v1/extract/folder", json={"folder_path": "/tmp"})
    assert resp.status_code == 401

    # Wrong key -> 401
    resp = client.post(
        "/api/v1/extract/folder",
        json={"folder_path": "/tmp"},
        headers={"X-API-Key": "wrong"},
    )
    assert resp.status_code == 401

    # Correct key -> passes auth (subsequent allowlist check yields 400, not 401)
    resp = client.post(
        "/api/v1/extract/folder",
        json={"folder_path": "/tmp"},
        headers={"X-API-Key": "s3cret"},
    )
    assert resp.status_code != 401
