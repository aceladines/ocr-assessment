import asyncio
import logging
from pathlib import Path
from typing import Any, Dict
from datetime import date, datetime
from fastapi.concurrency import run_in_threadpool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.exceptions import ExcelWriteError

logger = logging.getLogger("app.services.excel")

# Target columns for Excel output
HEADERS = [
    "Page",
    "Receipt No.",
    "Doctor Name",
    "PRC License",
    "Hospital",
    "Date",
    "Patient Name",
    "Total Amount (PHP)",
]


class ExcelService:
    """Service to build, style, and thread-safely update the extraction Excel sheets."""

    def __init__(self, output_dir: Path = settings.output_path):
        self.output_dir = output_dir
        self.lock = (
            asyncio.Lock()
        )  # Lock to ensure concurrent requests write sequentially

        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)

    async def write_batch_records(
        self, file_path: Path, records: list[Dict[str, Any]]
    ) -> None:
        """
        Safely writes and formats a sorted batch of records to the Excel file in a single pass.
        Builds the workbook fresh with styled headers. The blocking openpyxl work
        runs in a thread so the event loop is free.
        """
        wb_path = (self.output_dir / file_path.name).with_suffix(".xlsx")
        async with self.lock:
            try:
                logger.info(
                    f"Writing structured batch of {len(records)} records to Excel: {wb_path.name}"
                )
                await run_in_threadpool(self._write_batch_sync, wb_path, records)
                logger.info(
                    f"Successfully saved and formatted batch records to Excel: {wb_path.name}"
                )
            except Exception as exc:
                logger.error(
                    f"Error writing batch records to Excel at {wb_path}: {str(exc)}",
                    exc_info=True,
                )
                raise ExcelWriteError(
                    f"Failed to write extracted data to Excel spreadsheet: {str(exc)}"
                )

    def _write_batch_sync(
        self, wb_path: Path, records: list[Dict[str, Any]]
    ) -> None:
        """Blocking openpyxl build/save for a batch. Run via run_in_threadpool."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Invoices"
        ws.views.sheetView[0].showGridLines = True
        ws.append(HEADERS)
        self._apply_header_styling(ws)

        # Sort records sequentially by page number to guarantee perfect ordering
        sorted_records = sorted(records, key=lambda x: x.get("page", 1))
        for record in sorted_records:
            ws.append(self._map_record_to_row(record))
            self._apply_row_formatting(ws, ws.max_row)

        self._auto_fit_columns(ws)
        wb.save(wb_path)

    def _map_record_to_row(self, record: Dict[str, Any]) -> list:
        """Map OCR extraction fields to list values in exact header order."""
        # Clean Date parsing if possible
        date_val = record.get("date")
        if isinstance(date_val, str) and date_val:
            try:
                # Attempt to convert YYYY-MM-DD string to python date object
                date_val = datetime.strptime(date_val, "%Y-%m-%d").date()
            except ValueError:
                # Keep as original string if parsing fails
                pass

        return [
            record.get("page", 1),
            record.get("receipt_no") or "",
            record.get("doctor_name") or "",
            record.get("prc_license") or "",
            record.get("hospital") or "",
            date_val or "",
            record.get("patient_name") or "",
            record.get("total_amount") or 0.0,
        ]

    def _apply_header_styling(self, ws) -> None:
        """Style headers with professional Dark Blue background, white bold text, and nice borders."""
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="medium", color="1F4E78"),
            bottom=Side(style="medium", color="000000"),
        )

        ws.row_dimensions[1].height = 28

        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

    def _apply_row_formatting(self, ws, row_idx: int) -> None:
        """Apply cell borders, alignment, and native number/date formatting to data rows."""
        thin_border = Border(
            left=Side(style="thin", color="F2F2F2"),
            right=Side(style="thin", color="F2F2F2"),
            top=Side(style="thin", color="F2F2F2"),
            bottom=Side(style="thin", color="F2F2F2"),
        )
        data_font = Font(name="Calibri", size=11, bold=False, color="000000")

        ws.row_dimensions[row_idx].height = 20

        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border

            header_name = HEADERS[col_idx - 1]

            # Alignments & formats per column type
            if header_name == "Page":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0"
            elif header_name == "Date":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "yyyy-mm-dd"
            elif header_name == "Total Amount (PHP)":
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # Format as PHP currency: ₱#,##0.00
                cell.number_format = '"₱"#,##0.00'
            elif header_name in ["Receipt No.", "PRC License"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "@"  # Text format
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.number_format = "@"

    def _auto_fit_columns(self, ws) -> None:
        """Automatically adjust column widths to fit content with padding."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                # Convert cell value to string and check length
                if cell.value is not None:
                    # Special handling for date formatting length
                    if isinstance(cell.value, (datetime, date)):
                        val_str = "2026-12-12"
                    # Format float length for currency display
                    elif isinstance(cell.value, float):
                        val_str = f"₱{cell.value:,.2f}"
                    else:
                        val_str = str(cell.value)

                    max_len = max(max_len, len(val_str))

            # Use max length plus padding, constraint to min/max sizes
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)
